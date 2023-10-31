from tkinter import *
from tkinter import ttk
import getpass
import os
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from datetime import *
import pandas as pd

#current time
def get_time():
    date_and_time = datetime.today()
    current_time = date_and_time.strftime("%H")
    return current_time

#current date
def get_date():
    date_and_time = datetime.today()
    current_date = date_and_time.strftime("%m/%d/%Y")
    return current_date

#getting user info for creating a xl file
username = getpass.getuser()
pth = (os.path.join('Users',username,'Desktop','Pick Task Records.xlsx'))

#main window setup
root = Tk()
root.title("Pick Task Calculator")
root.minsize(width=470, height=300)

#top tab info
tab = ttk.Notebook(root)
tab.grid(row=0, column=0, columnspan=5)
show_list = False

#frames within tabs
frame1 = Frame(tab)
frame2 = Frame(tab)
frame3 = Frame(tab)

#frame info
frame1.pack(fill="both", expand=1)
frame2.pack(fill="both", expand=1)
frame3.pack(fill="both", expand=1)

#addinf frames to tab
tab.add(frame1, text="Main")
tab.add(frame2, text="Questions")
tab.add(frame3, text="Details")


multi_list = []
multi_time = []

sioc_list = []
sioc_time = []

sing_list = []
sing_time = []


dic = {0:0,
       1:0,
       2:0,
       3:0,
       4:0,
       5:0,
       6:0,
       7:0,
       8:0,
       9:0,
       10:0,
       11:0,
       12:0,
       13:0,
       14:0,
       15:0,
       16:0,
       17:0,
       18:0,
       19:0,
       20:0,
       21:0,
       22:0,
       23:0,      
}


#switch logic to show the list
def show_list_switch():
    global show_list

    if show_list:
        show_list_b.config(bg="white", borderwidth=5)
        show_list = False
        shown_list.config(text="")
    else:
        show_list_b.config(bg="grey", borderwidth=1)
        show_list = True

#all frame1 functions
currently_doing = None
def on_click():
    if currently_doing == "multi":
        add_multi()
    elif currently_doing == "sioc":
        add_sioc()
    elif currently_doing == "singles":
        add_sing()

def missed_time_adding(pack_list, time_list):
    current_time = int(get_time())
    adding = add.get()
    submited = submit_e.get()

    if adding.isdigit() and submited.isdigit():
        if int(adding)>0 and int(submited)<current_time:
            dic[int(submited)] += int(adding)
            pack_list.append(int(adding))
            time_list.append(int(submited))
            add.delete(0, END)
            submit_e.delete(0, END)
        else:
            add.delete(0, END)
            submit_e.delete(0, END)
            add.insert(0, "Time NOT Aplicable")
    else:
        add.delete(0, END)
        submit_e.delete(0, END)
        add.insert(0, "Please add a number")

def submit_time():
    if currently_doing == "multi":
        missed_time_adding(multi_list, multi_time)
        update_f1_info(multi_info, "Multi", multi_list, "red")
    elif currently_doing == "sioc":
        missed_time_adding(sioc_list, sioc_time)
        update_f1_info(sioc_info, "SIOC", sioc_list, "green")
    elif currently_doing == "singles":
        missed_time_adding(sing_list, sing_time)
        update_f1_info(sing_info, "Singles", sing_list, "blue")


#updates the frame 1 labels
def update_f1_info(info_label, name, list, color):
    info_label.config(text=name+": "+ str(sum(list)))

    if len(list) > 0:
        latest_added.config(text=f"Latest Added: {name} [{list[-1]}]", fg=color)
    else:
        latest_added.config(text=f"Latest Added: {name} (NONE)", fg=color)

    if show_list:
        shown_list.config(text=f"{name} List: {list}", fg=color)

    total_amount_label.config(text="Total: "+str(sum(multi_list)+sum(sioc_list)+sum(sing_list)))

# checks the (time when packed) with the (packed) and compairs it to the list you want to delete from and if thay match it deletes it.
def delete_from_dic(packed_list, time_list):
    dic[time_list[-1]] -= packed_list[-1]


def add_multi():
    global currently_doing
    add.config(bg="#ffd2d2")
    submit_b.config(bg="#ffd2d2")
    submit_e.config(bg="#ffd2d2")
    adding = add.get()
    current_time = int(get_time())
    if adding.isdigit():
        if int(adding)>0:
            dic[current_time] += int(adding)
            multi_list.append(int(adding))
            multi_time.append(current_time)
            add.delete(0, END)
        else:
            add.delete(0, END)
            add.insert(0, "Please Input a number greater than 0")
    elif adding=="del" and len(multi_list)>0:
        delete_from_dic(multi_list, multi_time)
        del multi_list[-1]
        del multi_time[-1]
        add.delete(0, END)
    else:
        add.delete(0, END)
        add.insert(0, "Please add a number")

    update_f1_info(multi_info, "Multi", multi_list, "red")
    
    currently_doing = "multi"

def add_sioc():
    global currently_doing
    add.config(bg="#d9ffd2")
    submit_b.config(bg="#d9ffd2")
    submit_e.config(bg="#d9ffd2")
    adding = add.get()
    current_time = int(get_time())
    if adding.isdigit():
        if int(adding)>0:
            dic[current_time] += int(adding)
            sioc_list.append(int(adding))
            sioc_time.append(current_time)
            add.delete(0, END)
        else:
            add.delete(0, END)
            add.insert(0, "Please Input a number greater than 0")
    elif adding=="del" and len(sioc_list)>0:
        delete_from_dic(sioc_list, sioc_time)
        del sioc_list[-1]
        del sioc_time[-1]
        add.delete(0, END)
    else:
        add.delete(0, END)
        add.insert(0, "Please add a number")

    update_f1_info(sioc_info, "SIOC", sioc_list, "green")

    currently_doing = "sioc"

def add_sing():
    global currently_doing
    add.config(bg="#d2d6ff")
    submit_b.config(bg="#d2d6ff")
    submit_e.config(bg="#d2d6ff")
    adding = add.get()
    current_time = int(get_time())
    if adding.isdigit():
        if int(adding)>0:
            dic[current_time] += int(adding)
            sing_list.append(int(adding))
            sing_time.append(current_time)
            add.delete(0, END)
        else:
            add.delete(0, END)
            add.insert(0, "Please Input a number greater than 0")
    elif adding=="del" and len(sing_list)>0:
        delete_from_dic(sing_list, sing_time)
        del sing_list[-1]
        del sing_time[-1]
        add.delete(0, END)
    else:
        add.delete(0, END)
        add.insert(0, "Please add a number")

    update_f1_info(sing_info, "Singles", sing_list, "blue")

    currently_doing = "singles"

def f1_clear_func():
    add.delete(0, END) 
    submit_e.delete(0, END)


#all frame2 functions
def sioc_question():
    qe1 = q_entry1.get()
    qe2 = q_entry2.get()
    if qe1.isdigit() and qe2.isdigit():
        qe1 = int(qe1)
        qe2 = int(qe2)
        layers = qe1/qe2
        product = qe1-(int(layers)*qe2)
        get1.config(text=f"Layers to get: {int(layers)}")
        get2.config(text=f"Products to get: {product}")
        q_entry2.delete(0,END)
    else:
        q_entry1.delete(0,END)
        q_entry2.delete(0,END)
        q_entry1.insert(0, 'Please input a number')

def singles_question():
    qe1 = q_entry1.get()
    qe2 = q_entry2.get()
    if qe1.isdigit() and qe2.isdigit():
        qe1 = int(qe1)
        qe2 = int(qe2)
        boxes = qe1/qe2
        deci_check = boxes-int(boxes)
        if deci_check > 0:
            products = deci_check*qe2
            get1.config(text=f"Boxes to get: {int(boxes)}")
            get2.config(text=f"Products to get: {round(products)}")
            q_entry2.delete(0,END)
        else:
            get1.config(text=f"Boxes to get: {int(boxes)}")
            get2.config(text="Products to get: 0")
            q_entry2.delete(0,END)
    else:
        q_entry1.delete(0,END)
        q_entry2.delete(0,END)
        q_entry1.insert(0, 'Please input a number')

def clear_func():
    q_entry1.delete(0,END)
    q_entry2.delete(0,END)
    get1.config(text=f"Layers to get: ")
    get2.config(text=f"Products to get: ")


#all frame3 functions
def update_func():
    title_f3.config(text=f"Total: {sum(multi_list)+sum(sioc_list)+sum(sing_list)}")
    multi_total_f3.config(text=f"Multi: {sum(multi_list)}")
    sioc_total_f3.config(text=f"SIOC: {sum(sioc_list)}")
    sing_total_f3.config(text=f"Singles: {sum(sing_list)}")

    pk_title_f3.config(text=f"Total Pick Tasks: {len(multi_list)+len(sioc_list)+len(sing_list)}")
    pk_multi_total_f3.config(text=f"Multi Pick Task: {len(multi_list)}")
    pk_sioc_total_f3.config(text=f"SIOC Pick Task: {len(sioc_list)}")
    pk_sing_total_f3.config(text=f"Singles Pick Task: {len(sing_list)}")
    check_func()
    
def check_func():
    save_b.config(state="active")
    if os.path.exists('C:\\'+pth):
        update_info.config(text="EXCEL File FOUND", fg="green")
    else:
        update_info.config(text="EXCEL File NOT FOUND", fg="red")


def save_func():
    t = sum(multi_list)+sum(sioc_list)+sum(sing_list)
    from datetime import date
    today = date.today()
    from openpyxl import load_workbook
    wb = load_workbook('C:\\'+pth)
    ws = wb.active
    ws.append([today, sum(multi_list), sum(sioc_list), sum(sing_list), t])
    wb.save('C:\\'+pth)
    save_b.config(bg="Green", state=DISABLED)

#creating a new xl file
def create_func():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Record"
    ws.append(["Date", "Multi", "SIOC", "Singles", "Total"])
    for col in range(1,6):
        ws[get_column_letter(col)+'1'].font=Font(bold=True)
    wb.save('C:\\'+pth)

#checks is a the xl file exists is not creates it and save the total
def save_check():
    if not os.path.exists('C:\\'+pth):
        create_func()
        save_func()
        update_info.config(text="Successfuly Created and Saved EXCEL File", fg="green")
    else:
        save_func()
        update_info.config(text="Successfuly Saved on EXCEL File", fg="green")


#open the graph window logic
graph_opened = False
def open_graph():
    last_time_added = 23
    while True:
        if dic[last_time_added]>0:
            break
        else:
            last_time_added -= 1
    graph_info = {'Time (24H Format)': [],
                  'Units Packed': []}
    first_time_added = True
    for k in dic.keys():
        if dic[k] == 0 and first_time_added:
            continue
        graph_info['Time (24H Format)'].append(k)
        graph_info['Units Packed'].append(dic[k])
        first_time_added = False
        if not first_time_added and k == last_time_added:
            break
    global canvas, top, graph_opened, df
    df = pd.DataFrame(graph_info)
    top = Toplevel()
    top.minsize(width=470, height=300)
    top.title("Graph "+get_date())
    figure1 = plt.Figure(figsize=(6, 5), dpi=100)
    ax1 = figure1.add_subplot(111)
    canvas = FigureCanvasTkAgg(figure1, top)
    canvas.get_tk_widget().pack()
    df = df[['Time (24H Format)', 'Units Packed']].groupby('Time (24H Format)').sum()
    df.plot(kind='bar', legend=True, ax=ax1)


    graph_opened = True


def open_and_update():
    if len(multi_list) > 0 or len(sioc_list) > 0 or len(sing_list) > 0:
        if graph_opened:
            top.destroy()
        open_graph()


#frame 1 (main logic and grid stats)
#entry for adding to list
add = Entry(frame1, width=55, borderwidth=5)
add.bind("<Return>", (lambda event: on_click()))
#grid info for the first text box
add.grid(row=1, column=0, columnspan=3, padx=10, pady=10)

#submit entry
submit_e = Entry(frame1, width=10, borderwidth=5)

submit_e.grid(row=4, column=4)

#all the adding buttons
multi_b = Button(frame1, text="MULTI", fg="Red", width=10, borderwidth=5, bg="#ffd2d2", command=add_multi)
sioc_b = Button(frame1, text="SIOC", fg="Green", width=10, borderwidth=5, bg="#d9ffd2", command=add_sioc)
sing_b = Button(frame1, text="SINGLES", fg="Blue", width=10, borderwidth=5, bg="#d2d6ff", command=add_sing)
#clear button for frame1
f1_clear_b = Button(frame1, text="Clear", borderwidth=5, width=5, height=1, bg="Grey", command=f1_clear_func)
#show list button
show_list_b = Button(frame1, text="Show List", borderwidth=5, width=7, height=1, bg="white", command=show_list_switch)
#missed time button
submit_b = Button(frame1, text="SUBMIT", borderwidth=5, width=6, height=1, bg="Grey", command=submit_time)

#current info labels
font_size = 12
total_amount_label = Label(frame1, text="Total: 0", font=("bold", 15))
multi_info = Label(frame1, text="Multi: 0", fg="Red", font=("bold",font_size))
sioc_info = Label(frame1, text="SIOC: 0", fg="Green", font=("bold",font_size))
sing_info = Label(frame1, text="Singles: 0", fg="Blue", font=("bold",font_size))

latest_added = Label(frame1, text="Latest Added: ")
shown_list = Label(frame1, text="")

#command info labels
cmnd = Label(frame1, text="Commands:")
del_cmnd = Label(frame1, text="[del] to Delete")


#grid info for the buttons
multi_b.grid(row=2, column=0, padx=5, pady=10)
sioc_b.grid(row=2, column=1, padx=5, pady=10)
sing_b.grid(row=2, column=2, padx=5, pady=10)
f1_clear_b.grid(row=3, column=0)
show_list_b.grid(row=3, column=2)
submit_b.grid(row=5, column=4)

#grid info for the current labels
total_amount_label.grid(row=5, column=2)
multi_info.grid(row=4, column=0, padx=10, pady=25)
sioc_info.grid(row=4, column=1, padx=10, pady=25)
sing_info.grid(row=4, column=2, padx=10, pady=25)

latest_added.grid(row=5, column=0, padx=10, pady=10)
shown_list.grid(row=7, column=0, columnspan=4, padx=10, pady=10)


#grid info for command labels
cmnd.grid(row=1, column=4)
del_cmnd.grid(row=2, column=4)


#question/frame2 system/logic
sioc_q = Button(frame2, text="SIOC Question", borderwidth=5, width=13, height=3, fg="Green", bg="#d9ffd2", command=sioc_question)
sing_q = Button(frame2, text="Singles Question", borderwidth=5, width=13, height=3, fg="Blue", bg="#d2d6ff", command=singles_question)

f2_clear_b = Button(frame2, text="Clear", borderwidth=5, width=7, height=2, bg="grey", command=clear_func)

q_entry1 = Entry(frame2, width=20, borderwidth=5)
q_entry2 = Entry(frame2, width=20, borderwidth=5)

entry_info1 = Label(frame2, text="Insert amount of products needed")
entry_info2 = Label(frame2, text="SIOC: Products per Layers / SINGLES: Products in Box")

get1 = Label(frame2, text="Layers to get: ")
get2 = Label(frame2, text="Products to get: ")

#questions/frame2 grid info
sioc_q.grid(row=0, column=0, padx=37, pady=5)
sing_q.grid(row=0, column=1, padx=37, pady=5)

f2_clear_b.grid(row=3, column=1, padx=37, pady=5)

q_entry1.grid(row=1, column=1, columnspan=3, pady=5, padx=5)
q_entry2.grid(row=2, column=1, columnspan=3, pady=5, padx=5)

entry_info1.grid(row=1, column=0)
entry_info2.grid(row=2, column=0)

get1.grid(row=3, column=0)
get2.grid(row=4, column=0)


#details/frame3 system/logic
update_b = Button(frame3, text="UPDATE", borderwidth=5, width=13, height=3, bg="grey", command=update_func)
save_b = Button(frame3, text="SAVE", borderwidth=5, width=13, height=3, bg="grey", state=DISABLED, command=save_check)
graph_button = Button(frame3, text="Open Graph", borderwidth=5, width=10, height=1, command=open_and_update)

title_f3 = Label(frame3, text=f"Total: {sum(multi_list)+sum(sioc_list)+sum(sing_list)}")
multi_total_f3 = Label(frame3, text=f"Multi: {sum(multi_list)}", fg="Red")
sioc_total_f3 = Label(frame3, text=f"SIOC: {sum(sioc_list)}", fg="Green")
sing_total_f3 = Label(frame3, text=f"Singles: {sum(sing_list)}", fg="Blue")

pk_title_f3 = Label(frame3, text=f"Total Pick Tasks: {len(multi_list)+len(sioc_list)+len(sing_list)}")
pk_multi_total_f3 = Label(frame3, text=f"Multi Pick Task: {len(multi_list)}", fg="Red")
pk_sioc_total_f3 = Label(frame3, text=f"SIOC Pick Task: {len(sioc_list)}", fg="Green")
pk_sing_total_f3 = Label(frame3, text=f"Singles Pick Task: {len(sing_list)}", fg="Blue")

update_info = Label(frame3, text=">>>Press Update<<<")

#details/frame3 grid info
padx_val = 45
title_f3.grid(row=0, column=0, padx=padx_val, pady=5)
multi_total_f3.grid(row=1, column=0, padx=padx_val)
sioc_total_f3.grid(row=2, column=0, padx=padx_val)
sing_total_f3.grid(row=3, column=0, padx=padx_val)

pk_title_f3.grid(row=0, column=1, padx=padx_val, pady=5)
pk_multi_total_f3.grid(row=1, column=1, padx=padx_val)
pk_sioc_total_f3.grid(row=2, column=1, padx=padx_val)
pk_sing_total_f3.grid(row=3, column=1, padx=padx_val)

update_info.grid(row=5, column=0)

update_b.grid(row=4, column=0, padx=padx_val, pady=30)
save_b.grid(row=4, column=1, padx=padx_val, pady=30)
graph_button.grid(row=5, column=1)


root.mainloop()
