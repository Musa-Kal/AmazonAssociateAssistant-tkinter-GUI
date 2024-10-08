from tkinter import *
from tkinter import ttk
import getpass
import os
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from datetime import *
import pandas as pd

#current time
def getCurrentHour():
    return datetime.today().strftime("%H")

#current date
def getCurrentDate():
    return datetime.today().strftime("%m/%d/%Y")

#getting user info for creating a xl file
pathToExcel = (os.path.join('Users',getpass.getuser(),'Desktop','Pick Task Records.xlsx'))

#main window setup
root = Tk()
root.title("Pick Task Calculator")
root.minsize(width=470, height=300)

#top tab info
tab = ttk.Notebook(root)
tab.grid(row=0, column=0, columnspan=5)
show_list = False

#frames within tabs
frame1 = Frame(tab)
frame2 = Frame(tab)
frame3 = Frame(tab)

#frame info
frame1.pack(fill="both", expand=1)
frame2.pack(fill="both", expand=1)
frame3.pack(fill="both", expand=1)

#addinf frames to tab
tab.add(frame1, text="Main")
tab.add(frame2, text="Get Pick Amount")
tab.add(frame3, text="Details")


amountPackedByTime = [0] * 24

class Task:
    def __init__(self, name: str, color1: str, color2: str) -> None:
        self.name = name
        self.color1 = color1
        self.colro2 = color2
        self.submitted = []
        self.timeSubmitted = []

SIOC = Task("sioc", "#d9ffd2", "green")
SINGLE = Task("singles", "#d2d6ff", "blue")
MULTI = Task("multi", "#ffd2d2", "red")


#switch logic to show the list
def show_list_switch():
    global show_list

    if show_list:
        show_list_b.config(bg="white", borderwidth=5)
        show_list = False
        shown_list.config(text="")
    else:
        show_list_b.config(bg="grey", borderwidth=1)
        show_list = True

#all frame1 functions
currently_doing = None
def on_click():
    if currently_doing == "multi":
        addToTask(MULTI)
    elif currently_doing == "sioc":
        addToTask(SIOC)
    elif currently_doing == "singles":
        addToTask(SINGLE)

def missed_time_adding(pack_list, time_list):
    current_time = int(getCurrentHour())
    adding = add.get()
    submited = submit_e.get()

    if adding.isdigit() and submited.isdigit():
        if int(adding)>0 and int(submited)<current_time:
            amountPackedByTime[int(submited)] += int(adding)
            pack_list.append(int(adding))
            time_list.append(int(submited))
            add.delete(0, END)
            submit_e.delete(0, END)
        else:
            add.delete(0, END)
            submit_e.delete(0, END)
            add.insert(0, "Time NOT Aplicable")
    else:
        add.delete(0, END)
        submit_e.delete(0, END)
        add.insert(0, "Please add a number")

def submit_time():
    if currently_doing == "multi":
        missed_time_adding(MULTI.submitted, MULTI.timeSubmitted)
        update_f1_info(multi_info, "Multi", MULTI.submitted, "red")
    elif currently_doing == "sioc":
        missed_time_adding(SIOC.submitted, SIOC.timeSubmitted)
        update_f1_info(sioc_info, "SIOC", SIOC.submitted, "green")
    elif currently_doing == "singles":
        missed_time_adding(SINGLE.submitted, SINGLE.timeSubmitted)
        update_f1_info(sing_info, "Singles", SINGLE.submitted, "blue")


#updates the frame 1 labels
def update_f1_info(info_label, name, taskList, color):
    info_label.config(text=name+": "+ str(sum(taskList)))

    if len(taskList) > 0:
        latest_added.config(text=f"Latest Added: {name} [{taskList[-1]}]", fg=color)
    else:
        latest_added.config(text=f"Latest Added: {name} (NONE)", fg=color)

    if show_list:
        shown_list.config(text=f"{name} List: {taskList}", fg=color)

    total_amount_label.config(text="Total: "+str(sum(MULTI.submitted)+sum(SIOC.submitted)+sum(SINGLE.submitted)))

# checks the (time when packed) with the (packed) and compairs it to the list you want to delete from and if thay match it deletes it.
def delete_from_amountPackedByTime(amount, time):
    amountPackedByTime[time] -= amount

def addToTask(task: Task):
    global currently_doing
    add.config(bg=task.color1)
    submit_b.config(bg=task.color1)
    submit_e.config(bg=task.color1)
    adding = add.get()
    current_time = int(getCurrentHour())
    if adding.isdigit():
        if int(adding)>0:
            amountPackedByTime[current_time] += int(adding)
            task.submitted.append(int(adding))
            task.timeSubmitted.append(current_time)
            add.delete(0, END)
        else:
            add.delete(0, END)
            add.insert(0, "Please Input a number greater than 0")
    elif adding=="del" and task.submitted:
        delete_from_amountPackedByTime(task.submitted.pop(), task.timeSubmitted.pop())
        add.delete(0, END)
    else:
        add.delete(0, END)
        add.insert(0, "Please add a number")

    infoLabel = sioc_info
    if task.name == "singles":
        infoLabel = sing_info
    elif task.name == "multi":
        infoLabel = multi_info

    update_f1_info(infoLabel, task.name.upper(), task.submitted, task.colro2)

    currently_doing = task.name

def f1_clear_func():
    add.delete(0, END) 
    submit_e.delete(0, END)


#all frame2 functions
def typeOfAmountToPick(task: str):
    typeToGet = "Layers"
    if task == "singles":
        typeToGet = "Boxes"

    qe1 = q_entry1.get()
    qe2 = q_entry2.get()
    
    if qe1.isdigit() and qe2.isdigit():
        qe1 = int(qe1)
        qe2 = int(qe2)

        if qe1 >= qe2:
            group = qe1//qe2
            product = qe1-(group*qe2)
            get1.config(text=f"{typeToGet} to get: {group}")
            get2.config(text=f"Products to get: {product}")
            q_entry2.delete(0,END)
        else:
            q_entry1.delete(0,END)
            q_entry2.delete(0,END)
            q_entry1.insert(0, 'Amoun to pick must be greater then or equal layers or boxes')

        
    else:
        q_entry1.delete(0,END)
        q_entry2.delete(0,END)
        q_entry1.insert(0, 'Please input a number')


def clear_func():
    q_entry1.delete(0,END)
    q_entry2.delete(0,END)
    get1.config(text=f"Layers to get: ")
    get2.config(text=f"Products to get: ")


#all frame3 functions
def update_func():
    title_f3.config(text=f"Total: {sum(MULTI.submitted)+sum(SIOC.submitted)+sum(SINGLE.submitted)}")
    multi_total_f3.config(text=f"Multi: {sum(MULTI.submitted)}")
    sioc_total_f3.config(text=f"SIOC: {sum(SIOC.submitted)}")
    sing_total_f3.config(text=f"Singles: {sum(SINGLE.submitted)}")

    pk_title_f3.config(text=f"Total Pick Tasks: {len(MULTI.submitted)+len(SIOC.submitted)+len(SINGLE.submitted)}")
    pk_multi_total_f3.config(text=f"Multi Pick Task: {len(MULTI.submitted)}")
    pk_sioc_total_f3.config(text=f"SIOC Pick Task: {len(SIOC.submitted)}")
    pk_sing_total_f3.config(text=f"Singles Pick Task: {len(SINGLE.submitted)}")
    check_func()
    
def check_func():
    save_b.config(state="active")
    if os.path.exists('C:\\'+pathToExcel):
        update_info.config(text="EXCEL File FOUND", fg="green")
    else:
        update_info.config(text="EXCEL File NOT FOUND", fg="red")


def save_func():
    t = sum(MULTI.submitted)+sum(SIOC.submitted)+sum(SINGLE.submitted)
    from datetime import date
    today = date.today()
    from openpyxl import load_workbook
    wb = load_workbook('C:\\'+pathToExcel)
    ws = wb.active
    ws.append([today, sum(MULTI.submitted), sum(SIOC.submitted), sum(SINGLE.submitted), t])
    wb.save('C:\\'+pathToExcel)
    save_b.config(bg="Green", state=DISABLED)

#creating a new xl file
def create_func():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Record"
    ws.append(["Date", "Multi", "SIOC", "Singles", "Total Packed"])
    for col in range(1,6):
        ws[get_column_letter(col)+'1'].font=Font(bold=True)
    wb.save('C:\\'+pathToExcel)

#checks is a the xl file exists is not creates it and save the total
def save_check():
    if not os.path.exists('C:\\'+pathToExcel):
        create_func()
        update_info.config(text="Successfuly Created and Saved EXCEL File", fg="green")
    else:

        update_info.config(text="Successfuly Saved on EXCEL File", fg="green")

    try:
        save_func()
    except Exception:
        update_info.config(text="ERROR: file could not be saved :(", fg="red")


#open the graph window logic
graph_opened = False

def open_graph():
    last_time_added = 23
    for t in range(23,0,-1):
        if amountPackedByTime[t]>0:
            last_time_added = t
            break
            
    graph_info = {'Time (24H Format)': [],
                  'Units Packed': []}
    first_time_added = True
    for t in range(len(amountPackedByTime)):
        if amountPackedByTime[t] == 0 and first_time_added:
            continue
        graph_info['Time (24H Format)'].append(t)
        graph_info['Units Packed'].append(amountPackedByTime[t])
        first_time_added = False
        if not first_time_added and t == last_time_added:
            break
    global canvas, top, graph_opened, df
    df = pd.DataFrame(graph_info)
    top = Toplevel()
    top.minsize(width=470, height=300)
    top.title("Graph "+getCurrentDate())
    figure1 = plt.Figure(figsize=(6, 5), dpi=100)
    ax1 = figure1.add_subplot(111)
    canvas = FigureCanvasTkAgg(figure1, top)
    canvas.get_tk_widget().pack()
    df = df[['Time (24H Format)', 'Units Packed']].groupby('Time (24H Format)').sum()
    df.plot(kind='bar', legend=True, ax=ax1)


    graph_opened = True


def open_and_update():
    if len(MULTI.submitted) > 0 or len(SIOC.submitted) > 0 or len(SINGLE.submitted) > 0:
        if graph_opened:
            top.destroy()
        open_graph()


#frame 1 (main logic and grid stats)
#entry for adding to list
add = Entry(frame1, width=55, borderwidth=5)
add.bind("<Return>", (lambda event: on_click()))
#grid info for the first text box
add.grid(row=1, column=0, columnspan=3, padx=10, pady=10)

#submit entry
submit_e = Entry(frame1, width=10, borderwidth=5)

submit_e.grid(row=4, column=4)

#all the adding buttons
multi_b = Button(frame1, text="MULTI", fg="Red", width=10, borderwidth=5, bg="#ffd2d2", command=lambda: addToTask(MULTI))
sioc_b = Button(frame1, text="SIOC", fg="Green", width=10, borderwidth=5, bg="#d9ffd2", command=lambda: addToTask(SIOC))
sing_b = Button(frame1, text="SINGLES", fg="Blue", width=10, borderwidth=5, bg="#d2d6ff", command=lambda: addToTask(SINGLE))
#clear button for frame1
f1_clear_b = Button(frame1, text="Clear", borderwidth=5, width=5, height=1, bg="Grey", command=f1_clear_func)
#show list button
show_list_b = Button(frame1, text="Show List", borderwidth=5, width=7, height=1, bg="white", command=show_list_switch)
#missed time button
submit_b = Button(frame1, text="SUBMIT", borderwidth=5, width=6, height=1, bg="Grey", command=submit_time)

#current info labels
font_size = 12
total_amount_label = Label(frame1, text="Total: 0", font=("bold", 15))
multi_info = Label(frame1, text="Multi: 0", fg="Red", font=("bold",font_size))
sioc_info = Label(frame1, text="SIOC: 0", fg="Green", font=("bold",font_size))
sing_info = Label(frame1, text="Singles: 0", fg="Blue", font=("bold",font_size))

latest_added = Label(frame1, text="Latest Added: ")
shown_list = Label(frame1, text="")

#command info labels
cmnd = Label(frame1, text="Commands:")
del_cmnd = Label(frame1, text="[del] to Delete")


#grid info for the buttons
multi_b.grid(row=2, column=0, padx=5, pady=10)
sioc_b.grid(row=2, column=1, padx=5, pady=10)
sing_b.grid(row=2, column=2, padx=5, pady=10)
f1_clear_b.grid(row=3, column=0)
show_list_b.grid(row=3, column=2)
submit_b.grid(row=5, column=4)

#grid info for the current labels
total_amount_label.grid(row=5, column=2)
multi_info.grid(row=4, column=0, padx=10, pady=25)
sioc_info.grid(row=4, column=1, padx=10, pady=25)
sing_info.grid(row=4, column=2, padx=10, pady=25)

latest_added.grid(row=5, column=0, padx=10, pady=10)
shown_list.grid(row=7, column=0, columnspan=4, padx=10, pady=10)


#grid info for command labels
cmnd.grid(row=1, column=4)
del_cmnd.grid(row=2, column=4)


#question/frame2 system/logic
sioc_q = Button(frame2, text="SIOC Question", borderwidth=5, width=13, height=3, fg="Green", bg="#d9ffd2", command=lambda: typeOfAmountToPick("sioc"))
sing_q = Button(frame2, text="Singles Question", borderwidth=5, width=13, height=3, fg="Blue", bg="#d2d6ff", command=lambda: typeOfAmountToPick("singles"))

f2_clear_b = Button(frame2, text="Clear", borderwidth=5, width=7, height=2, bg="grey", command=clear_func)

q_entry1 = Entry(frame2, width=20, borderwidth=5)
q_entry2 = Entry(frame2, width=20, borderwidth=5)

entry_info1 = Label(frame2, text="Insert amount of products needed")
entry_info2 = Label(frame2, text="SIOC: Products per Layers / SINGLES: Products in Box")

get1 = Label(frame2, text="Layers to get: ")
get2 = Label(frame2, text="Products to get: ")

#questions/frame2 grid info
sioc_q.grid(row=0, column=0, padx=37, pady=5)
sing_q.grid(row=0, column=1, padx=37, pady=5)

f2_clear_b.grid(row=3, column=1, padx=37, pady=5)

q_entry1.grid(row=1, column=1, columnspan=3, pady=5, padx=5)
q_entry2.grid(row=2, column=1, columnspan=3, pady=5, padx=5)

entry_info1.grid(row=1, column=0)
entry_info2.grid(row=2, column=0)

get1.grid(row=3, column=0)
get2.grid(row=4, column=0)


#details/frame3 system/logic
update_b = Button(frame3, text="UPDATE", borderwidth=5, width=13, height=3, bg="grey", command=update_func)
save_b = Button(frame3, text="SAVE", borderwidth=5, width=13, height=3, bg="grey", state=DISABLED, command=save_check)
graph_button = Button(frame3, text="Open Graph", borderwidth=5, width=10, height=1, command=open_and_update)

title_f3 = Label(frame3, text=f"Total: {sum(MULTI.submitted)+sum(SIOC.submitted)+sum(SINGLE.submitted)}")
multi_total_f3 = Label(frame3, text=f"Multi: {sum(MULTI.submitted)}", fg="Red")
sioc_total_f3 = Label(frame3, text=f"SIOC: {sum(SIOC.submitted)}", fg="Green")
sing_total_f3 = Label(frame3, text=f"Singles: {sum(SINGLE.submitted)}", fg="Blue")

pk_title_f3 = Label(frame3, text=f"Total Pick Tasks: {len(MULTI.submitted)+len(SIOC.submitted)+len(SINGLE.submitted)}")
pk_multi_total_f3 = Label(frame3, text=f"Multi Pick Task: {len(MULTI.submitted)}", fg="Red")
pk_sioc_total_f3 = Label(frame3, text=f"SIOC Pick Task: {len(SIOC.submitted)}", fg="Green")
pk_sing_total_f3 = Label(frame3, text=f"Singles Pick Task: {len(SINGLE.submitted)}", fg="Blue")

update_info = Label(frame3, text=">>>Press Update<<<")

#details/frame3 grid info
padx_val = 45
title_f3.grid(row=0, column=0, padx=padx_val, pady=5)
multi_total_f3.grid(row=1, column=0, padx=padx_val)
sioc_total_f3.grid(row=2, column=0, padx=padx_val)
sing_total_f3.grid(row=3, column=0, padx=padx_val)

pk_title_f3.grid(row=0, column=1, padx=padx_val, pady=5)
pk_multi_total_f3.grid(row=1, column=1, padx=padx_val)
pk_sioc_total_f3.grid(row=2, column=1, padx=padx_val)
pk_sing_total_f3.grid(row=3, column=1, padx=padx_val)

update_info.grid(row=5, column=0)

update_b.grid(row=4, column=0, padx=padx_val, pady=30)
save_b.grid(row=4, column=1, padx=padx_val, pady=30)
graph_button.grid(row=5, column=1)


root.mainloop()
