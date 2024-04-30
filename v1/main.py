from tkinter import *
from tkinter import ttk
import getpass
import os

#getting user info for creating a xl file
username = getpass.getuser()
pth = (os.path.join('Users',username,'Desktop','Pick Task Records.xlsx'))

#main window setup
root = Tk()
root.title("Pick Task Calculator")
root.minsize(width=470, height=300)

#top tab info
tab = ttk.Notebook(root)
tab.grid(row=0, column=0, columnspan=5)
show_list = False

#frames within tabs
frame1 = Frame(tab)
frame2 = Frame(tab)
frame3 = Frame(tab)

#frame info
frame1.pack(fill="both", expand=1)
frame2.pack(fill="both", expand=1)
frame3.pack(fill="both", expand=1)

#addinf frames to tab
tab.add(frame1, text="Main")
tab.add(frame2, text="Questions")
tab.add(frame3, text="Details")

sing_list = []
sioc_list = []
multi_list = []


#switch logic to show the list
def show_list_switch():
    global show_list

    if show_list:
        show_list_b.config(bg="white", borderwidth=5)
        show_list = False
        shown_list.config(text="")
    else:
        show_list_b.config(bg="grey", borderwidth=1)
        show_list = True

#all frame1 functions
currently_doing = None
def on_click():
    if currently_doing == "multi":
        add_multi()
    elif currently_doing == "sioc":
        add_sioc()
    elif currently_doing == "sing":
        add_sing()

def update_total_amount():
    total_amount_label.config(text="Total: "+str(sum(multi_list)+sum(sioc_list)+sum(sing_list)))

def add_multi():
    global currently_doing
    currently_doing = "multi"
    add.config(bg="#ffd2d2")

    adding = add.get()
    if adding.isdigit():
        if int(adding)>0:
            multi_list.append(int(adding))
            add.delete(0, END)
            multi_info.config(text="Multi: "+ str(sum(multi_list)))
            update_total_amount()
        else:
            add.delete(0, END)
            add.insert(0, "Please Input a number greater than 0")
    elif adding=="del" and sum(multi_list)>0:
        del multi_list[-1]
        add.delete(0, END)
        multi_info.config(text="Multi: "+ str(sum(multi_list)))
        update_total_amount()
    else:
        add.delete(0, END)
        add.insert(0, "Please add a number")
    if len(multi_list) > 0:
        latest_added.config(text=f"Latest Added: Multi [{multi_list[-1]}]", fg='red')
    else:
        latest_added.config(text="Latest Added: Multi (NONE)", fg='red')
    if show_list:
        shown_list.config(text=f"Multi List: {multi_list}", fg="red")

def add_sioc():
    global currently_doing
    currently_doing = "sioc"
    add.config(bg="#d9ffd2")

    adding = add.get()
    if adding.isdigit():
        if int(adding)>0:
            sioc_list.append(int(adding))
            add.delete(0, END)
            sioc_info.config(text="SIOC: "+ str(sum(sioc_list)))
            update_total_amount()
        else:
            add.delete(0, END)
            add.insert(0, "Please Input a number greater than 0")
    elif adding=="del" and len(sioc_list)>0:
        del sioc_list[-1]
        add.delete(0, END)
        sioc_info.config(text="SIOC: "+ str(sum(sioc_list)))
        update_total_amount()
    else:
        add.delete(0, END)
        add.insert(0, "Please add a number")
    if len(sioc_list) > 0:
        latest_added.config(text=f"Latest Added: SIOC [{sioc_list[-1]}]", fg='green')
    else:
        latest_added.config(text="Latest Added: SIOC (NONE)", fg='green')
    if show_list:
        shown_list.config(text=f"SIOC List: {sioc_list}", fg="green")

def add_sing():
    global currently_doing
    currently_doing = "sing"
    add.config(bg="#d2d6ff")

    adding = add.get()
    if adding.isdigit():
        if int(adding)>0:
            sing_list.append(int(adding))
            add.delete(0, END)
            sing_info.config(text="Singles: "+ str(sum(sing_list)))
            update_total_amount()
        else:
            add.delete(0, END)
            add.insert(0, "Please Input a number greater than 0")
    elif adding=="del" and sum(sing_list)>0:
        del sing_list[-1]
        add.delete(0, END)
        sing_info.config(text="Singles: "+ str(sum(sing_list)))
    else:
        add.delete(0, END)
        add.insert(0, "Please add a number")
    if len(sing_list) > 0:
        latest_added.config(text=f"Latest Added: Singles [{sing_list[-1]}]", fg='blue')
    else:
        latest_added.config(text="Latest Added: Singles (NONE)", fg='blue')
    if show_list:
          shown_list.config(text=f"Singles List: {sing_list}", fg="blue")


def f1_clear_func():
    add.delete(0, END)


#all frame2 functions
def sioc_question():
    qe1 = q_entry1.get()
    qe2 = q_entry2.get()
    if qe1.isdigit() and qe2.isdigit():
        qe1 = int(qe1)
        qe2 = int(qe2)
        layers = qe1/qe2
        product = qe1-(int(layers)*qe2)
        get1.config(text=f"Layers to get: {int(layers)}")
        get2.config(text=f"Products to get: {product}")
        q_entry2.delete(0,END)
    else:
        q_entry1.delete(0,END)
        q_entry2.delete(0,END)
        q_entry1.insert(0, 'Please input a number')

def singles_question():
    qe1 = q_entry1.get()
    qe2 = q_entry2.get()
    if qe1.isdigit() and qe2.isdigit():
        qe1 = int(qe1)
        qe2 = int(qe2)
        boxes = qe1/qe2
        deci_check = boxes-int(boxes)
        if deci_check > 0:
            products = deci_check*qe2
            get1.config(text=f"Boxes to get: {int(boxes)}")
            get2.config(text=f"Products to get: {round(products)}")
            q_entry2.delete(0,END)
        else:
            get1.config(text=f"Boxes to get: {int(boxes)}")
            get2.config(text="Products to get: 0")
            q_entry2.delete(0,END)
    else:
        q_entry1.delete(0,END)
        q_entry2.delete(0,END)
        q_entry1.insert(0, 'Please input a number')

def clear_func():
    q_entry1.delete(0,END)
    q_entry2.delete(0,END)
    get1.config(text=f"Layers to get: ")
    get2.config(text=f"Products to get: ")


#all frame3 functions
def update_func():
    title_f3.config(text=f"Total: {sum(multi_list)+sum(sioc_list)+sum(sing_list)}")
    multi_total_f3.config(text=f"Multi: {sum(multi_list)}")
    sioc_total_f3.config(text=f"SIOC: {sum(sioc_list)}")
    sing_total_f3.config(text=f"Singles: {sum(sing_list)}")

    pk_title_f3.config(text=f"Total Pick Tasks: {len(multi_list)+len(sioc_list)+len(sing_list)}")
    pk_multi_total_f3.config(text=f"Multi Pick Task: {len(multi_list)}")
    pk_sioc_total_f3.config(text=f"SIOC Pick Task: {len(sioc_list)}")
    pk_sing_total_f3.config(text=f"Singles Pick Task: {len(sing_list)}")
    check_func()
    
def check_func():
    save_b.config(state="active")
    if os.path.exists('C:\\'+pth):
        update_info.config(text="EXCEL File FOUND", fg="green")
    else:
        update_info.config(text="EXCEL File NOT FOUND", fg="red")


def save_func():
    t = sum(multi_list)+sum(sioc_list)+sum(sing_list)
    from datetime import date
    today = date.today()
    from openpyxl import load_workbook
    wb = load_workbook('C:\\'+pth)
    ws = wb.active
    ws.append([today, sum(multi_list), sum(sioc_list), sum(sing_list), t])
    wb.save('C:\\'+pth)
    save_b.config(bg="Green", state=DISABLED)

#creating a new xl file
def create_func():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Record"
    ws.append(["Date", "Multi", "SIOC", "Singles", "Total"])
    for col in range(1,6):
        ws[get_column_letter(col)+'1'].font=Font(bold=True)
    wb.save('C:\\'+pth)

#checks is a the xl file exists is not creates it and save the total
def save_check():
    if not os.path.exists('C:\\'+pth):
        create_func()
        save_func()
        update_info.config(text="Successfuly Created and Saved EXCEL File", fg="green")
    else:
        save_func()
        update_info.config(text="Successfuly Saved on EXCEL File", fg="green")


#frame 1 (main logic and grid stats)
#entry for adding to list
add = Entry(frame1, width=55, borderwidth=5)
add.bind("<Return>", (lambda event: on_click()))

#all the adding buttons
multi_b = Button(frame1, text="MULTI", fg="Red", width=10, borderwidth=5, bg="#ffd2d2", command=add_multi)
sioc_b = Button(frame1, text="SIOC", fg="Green", width=10, borderwidth=5, bg="#d9ffd2", command=add_sioc)
sing_b = Button(frame1, text="SINGLES", fg="Blue", width=10, borderwidth=5, bg="#d2d6ff", command=add_sing)
#clear button for frame1
f1_clear_b = Button(frame1, text="Clear", borderwidth=5, width=5, height=1, bg="Grey", command=f1_clear_func)
#show list button
show_list_b = Button(frame1, text="Show List", borderwidth=5, width=7, height=1, bg="white", command=show_list_switch)

#current info labels
font_size = 12
total_amount_label = Label(frame1, text="Total: 0", font=("bold", 15))
multi_info = Label(frame1, text="Multi: 0", fg="Red", font=("bold",font_size))
sioc_info = Label(frame1, text="SIOC: 0", fg="Green", font=("bold",font_size))
sing_info = Label(frame1, text="Singles: 0", fg="Blue", font=("bold",font_size))

latest_added = Label(frame1, text="Latest Added: ")
shown_list = Label(frame1, text="")

#command info labels
cmnd = Label(frame1, text="Commands:")
del_cmnd = Label(frame1, text="[del] to Delete")

#grid info for the first text box
add.grid(row=1, column=0, columnspan=3, padx=10, pady=10)

#grid info for the buttons
multi_b.grid(row=2, column=0, padx=5, pady=10)
sioc_b.grid(row=2, column=1, padx=5, pady=10)
sing_b.grid(row=2, column=2, padx=5, pady=10)
f1_clear_b.grid(row=3, column=0)
show_list_b.grid(row=3, column=2)

#grid info for the current labels
total_amount_label.grid(row=5, column=2)
multi_info.grid(row=4, column=0, padx=10, pady=25)
sioc_info.grid(row=4, column=1, padx=10, pady=25)
sing_info.grid(row=4, column=2, padx=10, pady=25)

latest_added.grid(row=5, column=0, padx=10, pady=10)
shown_list.grid(row=7, column=0, columnspan=4, padx=10, pady=10)


#grid info for command labels
cmnd.grid(row=1, column=4)
del_cmnd.grid(row=2, column=4)


#question/frame2 system/logic
sioc_q = Button(frame2, text="SIOC Question", borderwidth=5, width=13, height=3, fg="Green", bg="#d9ffd2", command=sioc_question)
sing_q = Button(frame2, text="Singles Question", borderwidth=5, width=13, height=3, fg="Blue", bg="#d2d6ff", command=singles_question)

f2_clear_b = Button(frame2, text="Clear", borderwidth=5, width=7, height=2, bg="grey", command=clear_func)

q_entry1 = Entry(frame2, width=20, borderwidth=5)
q_entry2 = Entry(frame2, width=20, borderwidth=5)

entry_info1 = Label(frame2, text="Insert amount of products needed")
entry_info2 = Label(frame2, text="SIOC: Products per Layers / SINGLES: Products in Box")

get1 = Label(frame2, text="Layers to get: ")
get2 = Label(frame2, text="Products to get: ")

#questions/frame2 grid info
sioc_q.grid(row=0, column=0, padx=37, pady=5)
sing_q.grid(row=0, column=1, padx=37, pady=5)

f2_clear_b.grid(row=3, column=1, padx=37, pady=5)

q_entry1.grid(row=1, column=1, columnspan=3, pady=5, padx=5)
q_entry2.grid(row=2, column=1, columnspan=3, pady=5, padx=5)

entry_info1.grid(row=1, column=0)
entry_info2.grid(row=2, column=0)

get1.grid(row=3, column=0)
get2.grid(row=4, column=0)


#details/frame3 system/logic
update_b = Button(frame3, text="UPDATE", borderwidth=5, width=13, height=3, bg="grey", command=update_func)
save_b = Button(frame3, text="SAVE", borderwidth=5, width=13, height=3, bg="grey", state=DISABLED, command=save_check)

title_f3 = Label(frame3, text=f"Total: {sum(multi_list)+sum(sioc_list)+sum(sing_list)}")
multi_total_f3 = Label(frame3, text=f"Multi: {sum(multi_list)}", fg="Red")
sioc_total_f3 = Label(frame3, text=f"SIOC: {sum(sioc_list)}", fg="Green")
sing_total_f3 = Label(frame3, text=f"Singles: {sum(sing_list)}", fg="Blue")

pk_title_f3 = Label(frame3, text=f"Total Pick Tasks: {len(multi_list)+len(sioc_list)+len(sing_list)}")
pk_multi_total_f3 = Label(frame3, text=f"Multi Pick Task: {len(multi_list)}", fg="Red")
pk_sioc_total_f3 = Label(frame3, text=f"SIOC Pick Task: {len(sioc_list)}", fg="Green")
pk_sing_total_f3 = Label(frame3, text=f"Singles Pick Task: {len(sing_list)}", fg="Blue")

update_info = Label(frame3, text=">>>Press Update<<<")

#details/frame3 grid info
padx_val = 45
title_f3.grid(row=0, column=0, padx=padx_val, pady=5)
multi_total_f3.grid(row=1, column=0, padx=padx_val)
sioc_total_f3.grid(row=2, column=0, padx=padx_val)
sing_total_f3.grid(row=3, column=0, padx=padx_val)

pk_title_f3.grid(row=0, column=1, padx=padx_val, pady=5)
pk_multi_total_f3.grid(row=1, column=1, padx=padx_val)
pk_sioc_total_f3.grid(row=2, column=1, padx=padx_val)
pk_sing_total_f3.grid(row=3, column=1, padx=padx_val)

update_info.grid(row=5, column=0)

update_b.grid(row=4, column=0, padx=padx_val, pady=30)
save_b.grid(row=4, column=1, padx=padx_val, pady=30)


root.mainloop()